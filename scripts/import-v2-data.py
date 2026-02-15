#!/usr/bin/env python3
"""
GATHER v2 — Excel → SQL Import Script

Reads the historical Excel files and generates SQL to populate the v2 database.

Data flow:
  1. Global Alumni Directory → contacts (292 alumni from all programs/years)
  2. CPF 2025 Tracker → enrich contacts, create CPF 2025 cohort members,
     import events, attendance, curriculum, ad hoc lists
  3. GGF 2025 Tracker → enrich contacts, create GGF 2025 cohort members,
     import events, attendance, curriculum

Key principle: Everyone is a contact. Alumni are just contacts who were
fellows in a past cohort. CPF/GGF 2025 fellows are contacts linked to a
live cohort. Same person, different cohort_members rows.

Usage:
  python3 scripts/import-v2-data.py > v2-import.sql
  # Then paste v2-import.sql into Supabase SQL Editor
"""

import os
import sys
import re
import json
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("-- pip install openpyxl", file=sys.stderr)
    sys.exit(1)

import warnings
warnings.filterwarnings("ignore")

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
HISTORICAL_DIR = os.path.join(BASE_DIR, "historical data")
GLOBAL_FILE = os.path.join(HISTORICAL_DIR, "Global Alumni Network Directory for Staff.xlsx")
CPF_FILE = os.path.join(HISTORICAL_DIR, "2025 CPF TRACKER.xlsx")
GGF_FILE = os.path.join(HISTORICAL_DIR, "2025 GGF TRACKER.xlsx")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def esc(val):
    """Escape value for SQL string literal. Returns 'value' or NULL."""
    if val is None:
        return "NULL"
    s = str(val).strip()
    s = re.sub(r'[\u200b-\u200f\u202a-\u202e\ufeff]', '', s)
    if not s or s.lower() in ("none", "n/a", "-", "—"):
        return "NULL"
    s = s.replace("'", "''")
    return f"'{s}'"


def esc_date(val):
    if val is None:
        return "NULL"
    if isinstance(val, datetime):
        return f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'"
    if isinstance(val, date):
        return f"'{val.isoformat()}'"
    return "NULL"


def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    s = re.sub(r'[\u200b-\u200f\u202a-\u202e\ufeff]', '', s)
    if not s or s.lower() in ("none", "n/a", "-", "—"):
        return None
    return s


def clean_phone(val):
    s = clean(val)
    if not s:
        return None
    if re.match(r'^\d+\.0$', s):
        s = s.replace('.0', '')
    return s


def clean_url(val):
    s = clean(val)
    if not s:
        return None
    return s.strip("/").strip()


def parse_birthday(val, fmt="mdy"):
    """Parse a birthday → (date, hide_year) or (None, None)."""
    if val is None:
        return None, None
    if isinstance(val, datetime):
        d = val.date()
        hide = d.year <= 1900 or d.year == 2000 or d.year >= 2005
        return d, hide
    if isinstance(val, date):
        hide = val.year <= 1900 or val.year == 2000 or val.year >= 2005
        return val, hide
    s = clean(val)
    if not s:
        return None, None
    parts = re.split(r'[/\-.]', s)
    if len(parts) != 3:
        return None, None
    try:
        if fmt == "mdy":
            m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
        else:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 1900 if y > 30 else 2000
        dt = date(y, m, d)
        hide = y <= 1900 or y == 2000 or y >= 2005
        return dt, hide
    except (ValueError, OverflowError):
        return None, None


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def parse_global_alumni():
    """Parse Global Alumni Directory → list of contact dicts."""
    if not os.path.exists(GLOBAL_FILE):
        print("-- Warning: Global Alumni file not found", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(GLOBAL_FILE, data_only=True)
    contacts = []
    seen = set()

    # GLOBAL sheet
    ws = wb["GLOBAL"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 19:
            continue
        email = clean(vals[8])
        if not email or "@" not in email:
            continue
        email = email.lower()
        if email in seen:
            continue
        seen.add(email)

        cohort_str = clean(vals[1])  # e.g. "2019 CPF", "2021 Global"
        program = None
        cohort_year = None
        if cohort_str:
            # Extract program
            cs = cohort_str.upper()
            if "CPF" in cs or "CHICAGO" in cs:
                program = "CPF"
            elif "GLOBAL" in cs or "GGF" in cs:
                program = "GGF"
            elif "ESP" in cs or "SPANISH" in cs:
                program = "ESP"
            elif "DAR" in cs or "TANZANIA" in cs:
                program = "DAR"
            elif "MOS" in cs or "COLOMBIA" in cs:
                program = "MOS"
            # Extract year
            ym = re.search(r'(\d{4})', cohort_str)
            if ym:
                cohort_year = ym.group(1)

        bday, hide_year = parse_birthday(vals[18], "mdy")

        contacts.append({
            "first_name": clean(vals[2]),
            "last_name": clean(vals[3]),
            "email": email,
            "phone": clean_phone(vals[7]),
            "website": clean_url(vals[9]),
            "facebook": clean_url(vals[10]),
            "facebook_org": clean_url(vals[11]),
            "twitter": clean_url(vals[12]),
            "twitter_org": clean_url(vals[13]),
            "instagram": clean_url(vals[14]),
            "instagram_org": clean_url(vals[15]),
            "linkedin": clean_url(vals[16]),
            "program": program,
            "cohort": cohort_year,
            "birthday": bday,
            "hide_birthday_year": hide_year,
            "status": "Alumni",
            "_cohort_program": program,
            "_cohort_year": int(cohort_year) if cohort_year else None,
        })

    # BIRTHDAYS sheet — enrich existing
    ws = wb["BIRTHDAYS"]
    email_to_contact = {c["email"]: c for c in contacts}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 18:
            continue
        email = clean(vals[8])
        if not email or "@" not in email:
            continue
        email = email.lower()
        bday, hide = parse_birthday(vals[17])
        if bday and email in email_to_contact and not email_to_contact[email].get("birthday"):
            email_to_contact[email]["birthday"] = bday
            email_to_contact[email]["hide_birthday_year"] = hide

    wb.close()
    return contacts


def parse_cpf_directory():
    """Parse CPF 2025 DIRECTORY → list of contact dicts (to merge with alumni)."""
    if not os.path.exists(CPF_FILE):
        return []

    wb = openpyxl.load_workbook(CPF_FILE, data_only=True)
    ws = wb["DIRECTORY"]
    contacts = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 5:
            continue
        email = clean(vals[4])
        if not email or "@" not in email:
            continue

        first = clean(vals[1])
        last = clean(vals[2])
        if not first or not last:
            continue

        bday, hide = parse_birthday(vals[12] if len(vals) > 12 else None, "mdy")

        contacts.append({
            "first_name": first,
            "last_name": last,
            "email": email.lower(),
            "phone": clean_phone(vals[3]),
            "job_title": clean(vals[5]),
            "organization": clean(vals[6]),
            "city": clean(vals[8]) if len(vals) > 8 else None,
            "country": "USA",
            "gender": clean(vals[11]) if len(vals) > 11 else None,
            "birthday": bday,
            "hide_birthday_year": hide,
            "community_area": clean(vals[14]) if len(vals) > 14 else None,
            "biography": clean(vals[18]) if len(vals) > 18 else None,
            "program": "CPF",
            "cohort": "2025",
            "status": "Current",
            "_cohort_program": "CPF",
            "_cohort_year": 2025,
        })

    wb.close()
    return contacts


def parse_ggf_directory():
    """Parse GGF 2025 DIRECTORY → list of contact dicts."""
    if not os.path.exists(GGF_FILE):
        return []

    wb = openpyxl.load_workbook(GGF_FILE, data_only=True)
    ws = wb["DIRECTORY"]
    contacts = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 4:
            continue
        email = clean(vals[3])
        if not email or "@" not in email:
            continue

        first = clean(vals[1])
        last = clean(vals[2])
        if not first or not last:
            continue

        bio = clean(vals[13]) if len(vals) > 13 else None
        if not bio:
            bio = clean(vals[12]) if len(vals) > 12 else None

        bday, hide = parse_birthday(vals[29] if len(vals) > 29 else None, "dmy")

        contacts.append({
            "first_name": first,
            "last_name": last,
            "email": email.lower(),
            "phone": clean_phone(vals[4]),
            "job_title": clean(vals[6]) if len(vals) > 6 else None,
            "organization": clean(vals[7]) if len(vals) > 7 else None,
            "country": clean(vals[8]) if len(vals) > 8 else None,
            "city": clean(vals[9]) if len(vals) > 9 else None,
            "biography": bio,
            "languages": clean(vals[14]) if len(vals) > 14 else None,
            "birthday": bday,
            "hide_birthday_year": hide,
            "website": clean_url(vals[16]) if len(vals) > 16 else None,
            "facebook": clean_url(vals[17]) if len(vals) > 17 else None,
            "twitter": clean_url(vals[18]) if len(vals) > 18 else None,
            "instagram": clean_url(vals[19]) if len(vals) > 19 else None,
            "linkedin": clean_url(vals[20]) if len(vals) > 20 else None,
            "tiktok": clean_url(vals[21]) if len(vals) > 21 else None,
            "website_org": clean_url(vals[22]) if len(vals) > 22 else None,
            "facebook_org": clean_url(vals[23]) if len(vals) > 23 else None,
            "twitter_org": clean_url(vals[25]) if len(vals) > 25 else None,
            "instagram_org": clean_url(vals[26]) if len(vals) > 26 else None,
            "linkedin_org": clean_url(vals[27]) if len(vals) > 27 else None,
            "program": "GGF",
            "cohort": "2025",
            "status": "Current",
            "_cohort_program": "GGF",
            "_cohort_year": 2025,
        })

    wb.close()
    return contacts


def parse_cpf_facilitators():
    """Parse CPF FACILITATORS → list of staff contact dicts."""
    if not os.path.exists(CPF_FILE):
        return []

    wb = openpyxl.load_workbook(CPF_FILE, data_only=True)
    ws = wb["FACILITATORS"]
    contacts = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 5:
            continue
        first = clean(vals[1])
        last = clean(vals[2])
        if not first or not last:
            continue

        email = clean(vals[14]) if len(vals) > 14 else None
        if not email or "@" not in email:
            email = clean(vals[5]) if len(vals) > 5 else None
        if not email or "@" not in email:
            continue

        contacts.append({
            "first_name": first,
            "last_name": last,
            "email": email.lower(),
            "phone": clean_phone(vals[3]),
            "job_title": clean(vals[7]) if len(vals) > 7 else None,
            "organization": clean(vals[8]) if len(vals) > 8 else None,
            "city": clean(vals[10]) if len(vals) > 10 else None,
            "biography": clean(vals[13]) if len(vals) > 13 else None,
            "status": "Staff",
            "_role": "team",
            "_cohort_program": "CPF",
            "_cohort_year": 2025,
            "_cohort_role": "facilitator",
        })

    wb.close()
    return contacts


def parse_schedule(filename, sheet_name, program):
    """Parse event schedule → list of event dicts."""
    if not os.path.exists(filename):
        return []

    wb = openpyxl.load_workbook(filename, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    events = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 8:
            continue

        if program == "CPF":
            event_date = vals[7]
            time_str = clean(vals[8])
            title = clean(vals[9])
            description = clean(vals[10]) if len(vals) > 10 else None
            location = clean(vals[11]) if len(vals) > 11 else None
            facilitator = clean(vals[6])
            notes = clean(vals[13]) if len(vals) > 13 else None
            theme = clean(vals[3])
            event_type_raw = clean(vals[5])
        else:
            event_date = vals[8] if len(vals) > 8 else None
            time_str = clean(vals[9]) if len(vals) > 9 else None
            title = clean(vals[7])
            description = None
            location = clean(vals[10]) if len(vals) > 10 else None
            facilitator = None
            notes = clean(vals[13]) if len(vals) > 13 else None
            theme = None
            event_type_raw = clean(vals[5]) if len(vals) > 5 else None

        if not isinstance(event_date, (datetime, date)):
            continue
        if not title:
            title = theme or "Session"

        event_type = "workshop"
        if event_type_raw:
            t = event_type_raw.lower()
            if "social" in t:
                event_type = "social"
            elif "orient" in t:
                event_type = "orientation"
            elif "field" in t:
                event_type = "field_trip"
            elif "guest" in t or "speaker" in t:
                event_type = "guest_speaker"

        start_time = event_date
        if isinstance(start_time, date) and not isinstance(start_time, datetime):
            start_time = datetime.combine(start_time, datetime.min.time())

        if time_str:
            try:
                t = re.match(r'(\d{1,2}):?(\d{2})?\s*(AM|PM)?', time_str, re.IGNORECASE)
                if t:
                    h = int(t.group(1))
                    m = int(t.group(2) or 0)
                    ampm = (t.group(3) or "").upper()
                    if ampm == "PM" and h < 12:
                        h += 12
                    elif ampm == "AM" and h == 12:
                        h = 0
                    start_time = start_time.replace(hour=h, minute=m)
            except (ValueError, AttributeError):
                pass

        events.append({
            "title": title,
            "description": description,
            "location": location,
            "start_time": start_time,
            "facilitator": facilitator,
            "event_type": event_type,
            "notes": notes,
            "program": program,
        })

    wb.close()
    return events


def parse_attendance(filename, sheet_name, program):
    """Parse attendance grid → list of dicts."""
    if not os.path.exists(filename):
        return [], []

    wb = openpyxl.load_workbook(filename, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return [], []

    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]

    date_cols = []
    for i, h in enumerate(headers):
        if isinstance(h, (datetime, date)):
            d = h if isinstance(h, date) else h.date()
            date_cols.append((i, d))

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 3:
            continue
        first = clean(vals[1])
        last = clean(vals[2])
        if not first or not last:
            continue

        for col_idx, event_date in date_cols:
            if col_idx >= len(vals):
                continue
            cell = vals[col_idx]
            if cell is None:
                continue

            status = None
            if isinstance(cell, (int, float)):
                if cell == 1 or cell == 1.0:
                    status = "present"
                elif cell == 0:
                    status = "absent"
            else:
                s = str(cell).strip().upper()
                if s in ("P", "1", "1.0"):
                    status = "present"
                elif s == "L":
                    status = "late"
                elif s == "E":
                    status = "excused"
                elif s in ("A", "0"):
                    status = "absent"

            if status:
                records.append({
                    "first_name": first,
                    "last_name": last,
                    "event_date": event_date,
                    "status": status,
                    "program": program,
                })

    wb.close()
    return records, date_cols


def parse_curriculum_progress(filename, program):
    """Parse CH progress sheets → chapter items + completions."""
    if not os.path.exists(filename):
        return [], [], []

    wb = openpyxl.load_workbook(filename, data_only=True)

    # Parse OVERVIEW for chapter page counts
    overview_sheet = "OVERVIEW" if program == "CPF" else "Overview"
    chapters = []
    if overview_sheet in wb.sheetnames:
        ws = wb[overview_sheet]
        headers = [str(c.value or "") for c in ws[1]]
        for i, h in enumerate(headers):
            m = re.match(r'Chapter\s+(\d+)\s*\n?\s*[xX]\s+of\s+(\d+)', h)
            if m:
                chapters.append({
                    "number": int(m.group(1)),
                    "total_pages": int(m.group(2)),
                    "col_idx": i,
                })

    # Parse individual CH progress sheets
    items_seen = set()
    all_items = []
    completions = []

    for name in wb.sheetnames:
        m = re.match(r'CH\s*(\d+)\s*(?:PROGRESS|Progress)', name)
        if not m:
            continue
        ch_num = int(m.group(1))
        ws = wb[name]

        # Row 2 has actual item names
        row2 = []
        if ws.max_row > 1:
            row2 = [str(c.value or "").strip() for c in ws[2]]

        # Find task/discussion columns from row 1 headers
        row1 = [str(c.value or "").strip() for c in ws[1]]
        items = []
        for i, h in enumerate(row1):
            h2 = row2[i] if i < len(row2) else ""
            item_name = h2 if h2 else h
            if not item_name:
                continue
            if "TASK" in h.upper():
                items.append({"col": i, "type": "assignment", "name": item_name})
            elif "DISC" in h.upper() or "DISCUSSION" in h.upper():
                items.append({"col": i, "type": "discussion", "name": item_name})

        # Register unique items
        for item in items:
            key = (program, ch_num, item["name"])
            if key not in items_seen:
                items_seen.add(key)
                all_items.append({
                    "chapter": ch_num,
                    "type": item["type"],
                    "name": item["name"],
                    "program": program,
                })

        # Parse completions (row 3+ since row 2 is sub-headers)
        for row in ws.iter_rows(min_row=3, values_only=True):
            vals = list(row)
            if len(vals) < 3:
                continue
            first = clean(vals[1]) if len(vals) > 1 else None
            last = clean(vals[2]) if len(vals) > 2 else None
            if not first or not last:
                continue

            for item in items:
                idx = item["col"]
                if idx >= len(vals) or vals[idx] is None:
                    continue
                cell = vals[idx]
                completed = False
                if isinstance(cell, (int, float)):
                    completed = cell >= 1
                elif str(cell).strip() in ("1", "1.0", "Y", "y"):
                    completed = True

                if completed:
                    completions.append({
                        "first_name": first,
                        "last_name": last,
                        "chapter": ch_num,
                        "item_name": item["name"],
                        "program": program,
                    })

    wb.close()
    return chapters, all_items, completions


def parse_adhoc_lists():
    """Parse CPF Ad Hoc Lists sheet."""
    if not os.path.exists(CPF_FILE):
        return []

    wb = openpyxl.load_workbook(CPF_FILE, data_only=True)
    lists = []

    if "Ad Hoc Lists" in wb.sheetnames:
        ws = wb["Ad Hoc Lists"]
        headers = [str(c.value or "").strip() for c in ws[1]]

        fields = []
        for i in range(4, len(headers)):
            h = headers[i]
            if h:
                fields.append({
                    "col": i,
                    "label": h,
                    "key": re.sub(r'\W+', '_', h.lower()).strip('_'),
                })

        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            first = clean(vals[1]) if len(vals) > 1 else None
            last = clean(vals[2]) if len(vals) > 2 else None
            if not first or not last:
                continue
            data = {}
            for f in fields:
                if f["col"] < len(vals) and vals[f["col"]] is not None:
                    v = vals[f["col"]]
                    if isinstance(v, (datetime, date)):
                        data[f["key"]] = v.isoformat()
                    elif isinstance(v, float) and v == int(v):
                        data[f["key"]] = str(int(v))
                    else:
                        s = str(v).strip()
                        if s:
                            data[f["key"]] = s
            if data:
                entries.append({
                    "first_name": first,
                    "last_name": last,
                    "data": data,
                })

        if fields:
            lists.append({
                "name": "Ad Hoc Tracking",
                "program": "CPF",
                "fields": [{"key": f["key"], "label": f["label"], "type": "text"} for f in fields],
                "entries": entries,
            })

    wb.close()
    return lists


# ---------------------------------------------------------------------------
# Merge contacts
# ---------------------------------------------------------------------------

def merge_contacts(alumni, cpf, ggf, facilitators):
    """
    Merge all contact sources. Same email = same person.
    CPF/GGF data enriches alumni records. New people get added.
    Returns list of unique contacts with their cohort memberships.
    """
    by_email = {}

    # Start with alumni (broadest set)
    for c in alumni:
        email = c["email"]
        by_email[email] = dict(c)
        by_email[email]["_memberships"] = []
        if c.get("_cohort_program") and c.get("_cohort_year"):
            by_email[email]["_memberships"].append({
                "program": c["_cohort_program"],
                "year": c["_cohort_year"],
                "role": "fellow",
            })

    # Merge CPF 2025 fellows
    for c in cpf:
        email = c["email"]
        if email in by_email:
            # Enrich existing: fill empty fields, update status
            existing = by_email[email]
            for k, v in c.items():
                if k.startswith("_"):
                    continue
                if v is not None and not existing.get(k):
                    existing[k] = v
            # Update status to Current (they're in a live cohort)
            existing["status"] = "Current"
            existing["program"] = "CPF"
            existing["cohort"] = "2025"
        else:
            # New contact (not in alumni directory)
            by_email[email] = dict(c)
            by_email[email]["_memberships"] = []

        by_email[email]["_memberships"].append({
            "program": "CPF",
            "year": 2025,
            "role": "fellow",
        })

    # Merge GGF 2025 fellows
    for c in ggf:
        email = c["email"]
        if email in by_email:
            existing = by_email[email]
            for k, v in c.items():
                if k.startswith("_"):
                    continue
                if v is not None and not existing.get(k):
                    existing[k] = v
            existing["status"] = "Current"
            existing["program"] = "GGF"
            existing["cohort"] = "2025"
        else:
            by_email[email] = dict(c)
            by_email[email]["_memberships"] = []

        by_email[email]["_memberships"].append({
            "program": "GGF",
            "year": 2025,
            "role": "fellow",
        })

    # Merge facilitators
    for c in facilitators:
        email = c["email"]
        if email in by_email:
            existing = by_email[email]
            for k, v in c.items():
                if k.startswith("_"):
                    continue
                if v is not None and not existing.get(k):
                    existing[k] = v
        else:
            by_email[email] = dict(c)
            by_email[email]["_memberships"] = []

        by_email[email].setdefault("_role", c.get("_role", "fellow"))
        by_email[email]["_memberships"].append({
            "program": c.get("_cohort_program", "CPF"),
            "year": c.get("_cohort_year", 2025),
            "role": c.get("_cohort_role", "facilitator"),
        })

    return list(by_email.values())


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------

def generate_sql():
    out = []

    def emit(s):
        out.append(s)

    # ---- Parse everything ----
    print("-- Parsing Global Alumni Directory...", file=sys.stderr)
    alumni = parse_global_alumni()
    print(f"--   {len(alumni)} alumni contacts", file=sys.stderr)

    print("-- Parsing CPF 2025 Directory...", file=sys.stderr)
    cpf = parse_cpf_directory()
    print(f"--   {len(cpf)} CPF fellows", file=sys.stderr)

    print("-- Parsing GGF 2025 Directory...", file=sys.stderr)
    ggf = parse_ggf_directory()
    print(f"--   {len(ggf)} GGF fellows", file=sys.stderr)

    print("-- Parsing CPF Facilitators...", file=sys.stderr)
    facilitators = parse_cpf_facilitators()
    print(f"--   {len(facilitators)} facilitators", file=sys.stderr)

    contacts = merge_contacts(alumni, cpf, ggf, facilitators)
    print(f"-- Merged: {len(contacts)} unique contacts", file=sys.stderr)

    # ---- Cohorts ----
    emit("-- =============================================")
    emit("-- GATHER v2 — Data Import")
    emit("-- =============================================")
    emit("")
    emit("-- Fix cohorts: replace 2026 seeds with real data")
    emit("DELETE FROM cohort_members WHERE cohort_id IN (SELECT id FROM cohorts WHERE cohort_year = 2026);")
    emit("DELETE FROM cohorts WHERE cohort_year = 2026;")
    emit("")

    # Collect all unique program+year combos for archived cohorts
    cohort_set = set()
    for c in contacts:
        for m in c.get("_memberships", []):
            if m["program"] and m["year"]:
                cohort_set.add((m["program"], m["year"]))

    # Always include CPF/GGF 2025 as live
    cohort_set.add(("CPF", 2025))
    cohort_set.add(("GGF", 2025))

    program_names = {
        "CPF": ("Chicago Peace Fellows", "Chicago", "USA"),
        "GGF": ("Global Gather Fellows", "Global", "Global"),
        "ESP": ("Spanish Language Fellows", "Global", "Global"),
        "DAR": ("Dar es Salaam Fellows", "Dar es Salaam", "Tanzania"),
        "MOS": ("Mosquera Fellows", "Mosquera", "Colombia"),
    }

    emit("-- Insert cohorts (2025 = live, all others = archived)")
    for program, year in sorted(cohort_set):
        info = program_names.get(program, (f"{program} Fellows", "Unknown", "Unknown"))
        name = f"{info[0]} {year}"
        city = info[1]
        country = info[2]
        status = "live" if year == 2025 else "archived"
        emit(f"INSERT INTO cohorts (name, program, city, country, cohort_year, status)")
        emit(f"VALUES ({esc(name)}, '{program}', {esc(city)}, {esc(country)}, {year}, '{status}')")
        emit(f"ON CONFLICT DO NOTHING;")
    emit("")

    # ---- Contacts ----
    emit("-- =============================================")
    emit(f"-- Contacts ({len(contacts)} total)")
    emit("-- =============================================")
    emit("")

    CONTACT_FIELDS = [
        "first_name", "last_name", "email", "phone", "job_title", "organization",
        "city", "country", "biography", "program", "cohort", "status", "gender",
        "community_area", "languages", "birthday", "hide_birthday_year",
        "linkedin", "twitter", "instagram", "facebook", "website", "tiktok",
        "linkedin_org", "twitter_org", "instagram_org", "facebook_org", "website_org",
    ]

    for c in contacts:
        cols = []
        vals = []
        for field in CONTACT_FIELDS:
            val = c.get(field)
            if val is None:
                continue
            cols.append(field)
            if field == "birthday":
                vals.append(f"'{val.isoformat()}'")
            elif field == "hide_birthday_year":
                vals.append("true" if val else "false")
            else:
                vals.append(esc(val))

        emit(f"INSERT INTO contacts ({', '.join(cols)})")
        emit(f"VALUES ({', '.join(vals)})")
        emit(f"ON CONFLICT DO NOTHING;")
        emit("")

    # ---- Contact roles ----
    emit("-- =============================================")
    emit("-- Contact roles")
    emit("-- =============================================")
    emit("")

    for c in contacts:
        role = c.get("_role", "fellow")
        emit(f"INSERT INTO contact_roles (contact_id, role)")
        emit(f"SELECT id, '{role}' FROM contacts WHERE LOWER(email) = {esc(c['email'])}")
        emit(f"ON CONFLICT (contact_id, role) DO NOTHING;")
    emit("")

    # ---- Cohort members ----
    emit("-- =============================================")
    emit("-- Cohort members")
    emit("-- =============================================")
    emit("")

    for c in contacts:
        for m in c.get("_memberships", []):
            if not m["program"] or not m["year"]:
                continue
            emit(f"INSERT INTO cohort_members (contact_id, cohort_id, role)")
            emit(f"SELECT c.id, co.id, '{m['role']}'")
            emit(f"FROM contacts c, cohorts co")
            emit(f"WHERE LOWER(c.email) = {esc(c['email'])}")
            emit(f"AND co.program = '{m['program']}' AND co.cohort_year = {m['year']}")
            emit(f"ON CONFLICT (contact_id, cohort_id) DO NOTHING;")
    emit("")

    # ---- Events ----
    print("-- Parsing events...", file=sys.stderr)
    cpf_events = parse_schedule(CPF_FILE, "2025 Schedule", "CPF")
    ggf_events = parse_schedule(GGF_FILE, "Schedule", "GGF")
    all_events = cpf_events + ggf_events
    print(f"--   {len(cpf_events)} CPF + {len(ggf_events)} GGF = {len(all_events)} events", file=sys.stderr)

    emit("-- =============================================")
    emit(f"-- Events ({len(all_events)})")
    emit("-- =============================================")
    emit("")

    for e in all_events:
        emit(f"INSERT INTO events (cohort_id, title, description, location, start_time, facilitator, event_type, notes)")
        emit(f"SELECT co.id, {esc(e['title'])}, {esc(e.get('description'))}, {esc(e.get('location'))}, {esc_date(e.get('start_time'))}, {esc(e.get('facilitator'))}, {esc(e.get('event_type'))}, {esc(e.get('notes'))}")
        emit(f"FROM cohorts co WHERE co.program = '{e['program']}' AND co.cohort_year = 2025;")
    emit("")

    # ---- Attendance ----
    print("-- Parsing attendance...", file=sys.stderr)
    cpf_att, cpf_dates = parse_attendance(CPF_FILE, "ATTENDANCE", "CPF")
    ggf_att, ggf_dates = parse_attendance(GGF_FILE, "Attendance", "GGF")
    all_att = cpf_att + ggf_att
    print(f"--   {len(cpf_att)} CPF + {len(ggf_att)} GGF = {len(all_att)} records", file=sys.stderr)

    emit("-- =============================================")
    emit(f"-- Attendance ({len(all_att)} records)")
    emit("-- =============================================")
    emit("")

    for r in all_att:
        emit(f"INSERT INTO event_attendance (event_id, contact_id, status)")
        emit(f"SELECT e.id, c.id, '{r['status']}'")
        emit(f"FROM events e")
        emit(f"JOIN cohorts co ON e.cohort_id = co.id AND co.program = '{r['program']}' AND co.cohort_year = 2025")
        emit(f"CROSS JOIN contacts c")
        emit(f"WHERE LOWER(c.first_name) = {esc(r['first_name'].lower())} AND LOWER(c.last_name) = {esc(r['last_name'].lower())}")
        emit(f"AND e.start_time::date = '{r['event_date'].isoformat()}'")
        emit(f"LIMIT 1")
        emit(f"ON CONFLICT (event_id, contact_id) DO NOTHING;")
    emit("")

    # ---- Curriculum ----
    print("-- Parsing curriculum...", file=sys.stderr)
    cpf_ch, cpf_items, cpf_comp = parse_curriculum_progress(CPF_FILE, "CPF")
    ggf_ch, ggf_items, ggf_comp = parse_curriculum_progress(GGF_FILE, "GGF")
    print(f"--   CPF: {len(cpf_ch)} chapters, {len(cpf_items)} items, {len(cpf_comp)} completions", file=sys.stderr)
    print(f"--   GGF: {len(ggf_ch)} chapters, {len(ggf_items)} items, {len(ggf_comp)} completions", file=sys.stderr)

    emit("-- =============================================")
    emit("-- Curriculum")
    emit("-- =============================================")
    emit("")

    # Create curricula
    emit("INSERT INTO curricula (cohort_id, name, year)")
    emit("SELECT id, name || ' Curriculum', 2025 FROM cohorts WHERE cohort_year = 2025;")
    emit("")

    # Create chapters
    for program, chapters in [("CPF", cpf_ch), ("GGF", ggf_ch)]:
        for ch in chapters:
            emit(f"INSERT INTO curriculum_chapters (curriculum_id, title, description, order_num)")
            emit(f"SELECT cu.id, 'Chapter {ch['number']}', '{ch['total_pages']} pages', {ch['number']}")
            emit(f"FROM curricula cu JOIN cohorts co ON cu.cohort_id = co.id")
            emit(f"WHERE co.program = '{program}' AND co.cohort_year = 2025;")

            # Page items
            for i in range(1, ch["total_pages"] + 1):
                emit(f"INSERT INTO curriculum_items (chapter_id, item_type, title, order_num)")
                emit(f"SELECT cc.id, 'page', 'Page {i}', {i}")
                emit(f"FROM curriculum_chapters cc JOIN curricula cu ON cc.curriculum_id = cu.id")
                emit(f"JOIN cohorts co ON cu.cohort_id = co.id")
                emit(f"WHERE co.program = '{program}' AND co.cohort_year = 2025 AND cc.title = 'Chapter {ch['number']}';")
    emit("")

    # Task/discussion items
    for program, items in [("CPF", cpf_items), ("GGF", ggf_items)]:
        for item in items:
            emit(f"INSERT INTO curriculum_items (chapter_id, item_type, title, order_num)")
            emit(f"SELECT cc.id, '{item['type']}', {esc(item['name'])}, 100")
            emit(f"FROM curriculum_chapters cc JOIN curricula cu ON cc.curriculum_id = cu.id")
            emit(f"JOIN cohorts co ON cu.cohort_id = co.id")
            emit(f"WHERE co.program = '{program}' AND co.cohort_year = 2025 AND cc.title = 'Chapter {item['chapter']}';")
    emit("")

    # Completions
    emit("-- Curriculum completions")
    for program, comps in [("CPF", cpf_comp), ("GGF", ggf_comp)]:
        for comp in comps:
            emit(f"INSERT INTO contact_curriculum_progress (contact_id, item_id, completed, completed_at)")
            emit(f"SELECT c.id, ci.id, true, now()")
            emit(f"FROM contacts c, curriculum_items ci")
            emit(f"JOIN curriculum_chapters cc ON ci.chapter_id = cc.id")
            emit(f"JOIN curricula cu ON cc.curriculum_id = cu.id")
            emit(f"JOIN cohorts co ON cu.cohort_id = co.id")
            emit(f"WHERE LOWER(c.first_name) = {esc(comp['first_name'].lower())}")
            emit(f"AND LOWER(c.last_name) = {esc(comp['last_name'].lower())}")
            emit(f"AND co.program = '{program}' AND co.cohort_year = 2025")
            emit(f"AND cc.title = 'Chapter {comp['chapter']}' AND ci.title = {esc(comp['item_name'])}")
            emit(f"ON CONFLICT (contact_id, item_id) DO NOTHING;")
    emit("")

    # ---- Ad Hoc Lists ----
    print("-- Parsing ad hoc lists...", file=sys.stderr)
    adhoc = parse_adhoc_lists()
    print(f"--   {len(adhoc)} lists", file=sys.stderr)

    if adhoc:
        emit("-- =============================================")
        emit("-- Ad Hoc Lists")
        emit("-- =============================================")
        emit("")

        for lst in adhoc:
            fields_json = json.dumps(lst["fields"]).replace("'", "''")
            emit(f"INSERT INTO adhoc_lists (cohort_id, name, fields)")
            emit(f"SELECT co.id, {esc(lst['name'])}, '{fields_json}'::jsonb")
            emit(f"FROM cohorts co WHERE co.program = '{lst['program']}' AND co.cohort_year = 2025;")
            emit("")

            for entry in lst["entries"]:
                data_json = json.dumps(entry["data"]).replace("'", "''")
                emit(f"INSERT INTO adhoc_list_entries (list_id, contact_id, data)")
                emit(f"SELECT al.id, c.id, '{data_json}'::jsonb")
                emit(f"FROM adhoc_lists al")
                emit(f"JOIN cohorts co ON al.cohort_id = co.id AND co.program = '{lst['program']}'")
                emit(f"JOIN contacts c ON LOWER(c.first_name) = {esc(entry['first_name'].lower())} AND LOWER(c.last_name) = {esc(entry['last_name'].lower())}")
                emit(f"WHERE al.name = {esc(lst['name'])}")
                emit(f"ON CONFLICT (list_id, contact_id) DO NOTHING;")
        emit("")

    emit("-- =============================================")
    emit("-- DONE")
    emit("-- =============================================")

    return "\n".join(out)


if __name__ == "__main__":
    sql = generate_sql()
    print(sql)
