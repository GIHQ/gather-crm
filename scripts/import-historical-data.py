#!/usr/bin/env python3
"""
GATHER CRM - Historical Data Import Script

Parses the three historical Excel trackers and enriches existing Supabase
fellows records with birthdays, social links, gender, phone numbers, bios,
cohort data, and more.

Data sources:
  1. Global Alumni Network Directory for Staff.xlsx
     - GLOBAL sheet: 292 alumni with social links, birthdays, cohort, phone
     - BIRTHDAYS sheet: curated subset with birthday data
  2. 2025 CPF TRACKER.xlsx
     - DIRECTORY: 20 CPF fellows with DOB, gender, phone, bio, social
  3. 2025 GGF TRACKER.xlsx
     - DIRECTORY: ~25 GGF fellows with DOB, phone, social links, bio, languages

Matching strategy: Email-based (case-insensitive) against existing Supabase fellows.
Update strategy: Only fill empty/null fields — never overwrite existing data.

Environment variables required:
  SUPABASE_URL - Supabase project URL
  SUPABASE_SERVICE_ROLE_KEY - Service role key (bypasses RLS)

Usage:
  python3 scripts/import-historical-data.py              # Dry run (default)
  python3 scripts/import-historical-data.py --apply      # Actually write to DB
  python3 scripts/import-historical-data.py --verbose     # Extra detail
"""

import os
import sys
import json
import re
from datetime import datetime, date
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Error: requests required. Install with: pip install requests")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
HISTORICAL_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "historical data")
GLOBAL_ALUMNI_FILE = os.path.join(HISTORICAL_DIR, "Global Alumni Network Directory for Staff.xlsx")
CPF_TRACKER_FILE = os.path.join(HISTORICAL_DIR, "2025 CPF TRACKER.xlsx")
GGF_TRACKER_FILE = os.path.join(HISTORICAL_DIR, "2025 GGF TRACKER.xlsx")

# Fields that can be enriched (DB column → human label)
ENRICHABLE_FIELDS = {
    "birthday": "Birthday",
    "hide_birthday_year": "Hide Birthday Year",
    "gender": "Gender",
    "phone": "Phone",
    "linkedin": "LinkedIn",
    "twitter": "Twitter/X",
    "instagram": "Instagram",
    "facebook": "Facebook",
    "website": "Website",
    "linkedin_org": "LinkedIn (Org)",
    "twitter_org": "Twitter/X (Org)",
    "instagram_org": "Instagram (Org)",
    "facebook_org": "Facebook (Org)",
    "website_org": "Website (Org)",
    "tiktok": "TikTok",
    "biography": "Biography",
    "languages": "Languages",
    "cohort": "Cohort",
    "community_area": "Community Area",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(val):
    """Clean a cell value: strip whitespace, normalize None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return str(val).strip()
    s = str(val).strip().replace("\u202d", "").replace("\u202c", "")
    # Remove trailing/leading newlines
    s = s.strip()
    if not s or s.lower() in ("none", "n/a", "-", "—"):
        return None
    return s


def clean_url(val):
    """Clean and normalize a URL/social link."""
    s = clean(val)
    if not s:
        return None
    # Remove common junk
    s = s.strip("/").strip()
    if not s:
        return None
    return s


def clean_phone(val):
    """Clean a phone number — preserve international format."""
    s = clean(val)
    if not s:
        return None
    # Remove invisible unicode characters
    s = re.sub(r'[\u200b-\u200f\u202a-\u202e\ufeff]', '', s)
    # If it looks like a float (e.g., "17733648000.0"), clean it
    if re.match(r'^\d+\.0$', s):
        s = s.replace('.0', '')
    s = s.strip()
    if not s:
        return None
    return s


def parse_birthday_string(s, fmt="mdy"):
    """
    Parse a birthday string into (date, hide_year).
    fmt: 'mdy' for MM/DD/YYYY, 'dmy' for DD/MM/YYYY
    Returns (date_obj, hide_birthday_year) or (None, None).
    """
    s = clean(s)
    if not s:
        return None, None

    parts = re.split(r'[/\-.]', s)
    if len(parts) != 3:
        return None, None

    try:
        if fmt == "mdy":
            month, day, year = int(parts[0]), int(parts[1]), int(parts[2])
        else:  # dmy
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])

        # Handle 2-digit years
        if year < 100:
            year += 1900 if year > 30 else 2000

        # Validate ranges
        if month < 1 or month > 12 or day < 1 or day > 31:
            return None, None

        d = date(year, month, day)

        # Check if year seems like a placeholder
        hide_year = year <= 1900 or year >= 2005
        return d, hide_year

    except (ValueError, OverflowError):
        return None, None


def parse_birthday_datetime(dt, cohort_str=None):
    """
    Parse a datetime object from Excel into (date, hide_year).
    Handles suspicious year=2000 placeholder pattern.

    When Excel stores a date that was entered as only month/day (no year),
    it typically defaults to year 1900 or 2000. We detect this pattern
    and set hide_birthday_year=True so the UI only shows month/day.
    """
    if dt is None:
        return None, None

    if isinstance(dt, str):
        # Sometimes Excel gives strings even in date columns
        # Try both formats
        for fmt in ["mdy", "dmy"]:
            d, hide = parse_birthday_string(dt, fmt)
            if d:
                return d, hide
        return None, None

    if isinstance(dt, datetime):
        d = dt.date()
    elif isinstance(dt, date):
        d = dt
    else:
        return None, None

    year = d.year

    # Parse cohort year for sanity checking
    cohort_year = None
    if cohort_str:
        m = re.search(r'(\d{4})', str(cohort_str))
        if m:
            cohort_year = int(m.group(1))

    # Flag suspicious years
    hide_year = False
    if year <= 1900:
        # Excel default placeholder year
        hide_year = True
    elif year == 2000:
        # Very common Excel placeholder when only month/day was entered.
        # 24 of ~60 BIRTHDAYS sheet records have year=2000 — statistically
        # impossible for this to be real across a fellowship program.
        hide_year = True
    elif cohort_year and year > cohort_year:
        # Born after joining program — impossible
        hide_year = True
    elif cohort_year and (cohort_year - year) < 15:
        # Would have been under 15 at time of fellowship — very unlikely
        hide_year = True

    return d, hide_year


def extract_cohort_year(cohort_str):
    """Extract the year from a cohort string like '2019 CPF' or '2021 Global'."""
    if not cohort_str:
        return None
    m = re.search(r'(\d{4})', str(cohort_str))
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Excel Parsers
# ---------------------------------------------------------------------------

def parse_global_alumni():
    """Parse the Global Alumni Network Directory — GLOBAL + BIRTHDAYS sheets."""
    records = {}  # email → dict of fields

    if not os.path.exists(GLOBAL_ALUMNI_FILE):
        print(f"  Warning: {GLOBAL_ALUMNI_FILE} not found, skipping")
        return records

    wb = openpyxl.load_workbook(GLOBAL_ALUMNI_FILE, data_only=True)

    # --- GLOBAL sheet ---
    ws = wb["GLOBAL"]
    print(f"  GLOBAL sheet: {ws.max_row - 1} data rows")

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 19:
            continue

        email = clean(vals[8])
        if not email or "@" not in email:
            continue
        email = email.lower()

        cohort_str = clean(vals[1])
        cohort_year = extract_cohort_year(cohort_str)

        # Parse birthday (MM/DD/YYYY string format in this sheet)
        bday_raw = vals[18]
        bday, hide_year = None, None
        if isinstance(bday_raw, str):
            bday, hide_year = parse_birthday_string(bday_raw, fmt="mdy")
        elif isinstance(bday_raw, (datetime, date)):
            bday, hide_year = parse_birthday_datetime(bday_raw, cohort_str)

        rec = {
            "source": "Global Alumni Directory (GLOBAL)",
            "first_name": clean(vals[2]),
            "last_name": clean(vals[3]),
            "cohort": cohort_year,
            "phone": clean_phone(vals[7]),
            "website": clean_url(vals[9]),
            "facebook": clean_url(vals[10]),      # PERSONAL facebook
            "facebook_org": clean_url(vals[11]),   # ORGANIZATION facebook
            "twitter": clean_url(vals[12]),        # PERSONAL twitter
            "twitter_org": clean_url(vals[13]),    # ORGANIZATION twitter
            "instagram": clean_url(vals[14]),      # PERSONAL instagram
            "instagram_org": clean_url(vals[15]),  # ORGANIZATION instagram
            "linkedin": clean_url(vals[16]),       # LINKED IN
        }

        if bday:
            rec["birthday"] = bday.isoformat()
            rec["hide_birthday_year"] = hide_year

        # Only add non-None values
        records[email] = {k: v for k, v in rec.items() if v is not None}

    # --- BIRTHDAYS sheet (curated subset with more birthday data) ---
    ws = wb["BIRTHDAYS"]
    bday_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 18:
            continue

        email = clean(vals[8])
        if not email or "@" not in email:
            continue
        email = email.lower()

        cohort_str = clean(vals[1])
        bday_raw = vals[17]  # BIRTHDAY col in BIRTHDAYS sheet
        bday, hide_year = parse_birthday_datetime(bday_raw, cohort_str)

        if bday:
            if email not in records:
                records[email] = {"source": "Global Alumni Directory (BIRTHDAYS)"}
            # Birthday from BIRTHDAYS sheet — may override GLOBAL if GLOBAL was empty
            if "birthday" not in records[email]:
                records[email]["birthday"] = bday.isoformat()
                records[email]["hide_birthday_year"] = hide_year
                bday_count += 1

        # Also extract social links from BIRTHDAYS sheet
        social_fields = {
            "phone": clean_phone(vals[7]),
            "website": clean_url(vals[9]),
            "facebook": clean_url(vals[10]),
            "facebook_org": clean_url(vals[11]),
            "twitter": clean_url(vals[12]),
            "twitter_org": clean_url(vals[13]),
            "instagram": clean_url(vals[14]),
            "instagram_org": clean_url(vals[15]),
            "linkedin": clean_url(vals[16]),
        }

        if email in records:
            for k, v in social_fields.items():
                if v and k not in records[email]:
                    records[email][k] = v

    print(f"  BIRTHDAYS sheet: {bday_count} additional birthdays extracted")
    wb.close()
    return records


def parse_cpf_tracker():
    """Parse the 2025 CPF TRACKER — DIRECTORY sheet."""
    records = {}

    if not os.path.exists(CPF_TRACKER_FILE):
        print(f"  Warning: {CPF_TRACKER_FILE} not found, skipping")
        return records

    wb = openpyxl.load_workbook(CPF_TRACKER_FILE, data_only=True)
    ws = wb["DIRECTORY"]
    print(f"  CPF DIRECTORY: {ws.max_row - 1} data rows")

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 18:
            continue

        email = clean(vals[4])
        if not email or "@" not in email:
            continue
        email = email.lower()

        # Parse D.O.B. (mixed: MM/DD/YYYY strings or datetime objects)
        dob_raw = vals[12]
        bday, hide_year = None, None
        if isinstance(dob_raw, str):
            bday, hide_year = parse_birthday_string(dob_raw, fmt="mdy")
        elif isinstance(dob_raw, (datetime, date)):
            bday, hide_year = parse_birthday_datetime(dob_raw, "2025 CPF")

        rec = {
            "source": "2025 CPF Tracker",
            "first_name": clean(vals[1]),
            "last_name": clean(vals[2]),
            "phone": clean_phone(vals[3]),
            "gender": clean(vals[11]),
            "community_area": clean(vals[14]),
            "biography": clean(vals[18]) if len(vals) > 18 else None,
        }

        # Social links (Organization Social at col 16, Personal Social at col 17)
        org_social = clean_url(vals[16]) if len(vals) > 16 else None
        personal_social = clean_url(vals[17]) if len(vals) > 17 else None

        # Try to categorize social links by URL pattern
        for url, is_org in [(org_social, True), (personal_social, False)]:
            if url:
                categorized = categorize_social_url(url, is_org)
                rec.update(categorized)

        if bday:
            rec["birthday"] = bday.isoformat()
            rec["hide_birthday_year"] = hide_year

        records[email] = {k: v for k, v in rec.items() if v is not None}

    wb.close()
    return records


def parse_ggf_tracker():
    """Parse the 2025 GGF TRACKER — DIRECTORY sheet."""
    records = {}

    if not os.path.exists(GGF_TRACKER_FILE):
        print(f"  Warning: {GGF_TRACKER_FILE} not found, skipping")
        return records

    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    wb = openpyxl.load_workbook(GGF_TRACKER_FILE, data_only=True)
    ws = wb["DIRECTORY"]
    print(f"  GGF DIRECTORY: {ws.max_row - 1} data rows")

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 20:
            continue

        email = clean(vals[3])
        if not email or "@" not in email:
            continue
        email = email.lower()

        # Parse Date of Birth (DD/MM/YYYY format per header)
        dob_raw = vals[29] if len(vals) > 29 else None
        bday, hide_year = None, None
        if isinstance(dob_raw, str):
            dob_cleaned = dob_raw.strip()
            bday, hide_year = parse_birthday_string(dob_cleaned, fmt="dmy")
        elif isinstance(dob_raw, (datetime, date)):
            bday, hide_year = parse_birthday_datetime(dob_raw, "2025 GGF")

        # Use updated bio if available, otherwise original bio
        bio = clean(vals[13]) or clean(vals[12])  # UPDATED BIO or BIO

        rec = {
            "source": "2025 GGF Tracker",
            "first_name": clean(vals[1]),
            "last_name": clean(vals[2]),
            "phone": clean_phone(vals[4]) or clean_phone(vals[5]),  # Cell/SMS or WhatsApp
            "biography": bio,
            "languages": clean(vals[14]) if len(vals) > 14 else None,
            "website": clean_url(vals[16]) if len(vals) > 16 else None,
            "facebook": clean_url(vals[17]) if len(vals) > 17 else None,
            "twitter": clean_url(vals[18]) if len(vals) > 18 else None,
            "instagram": clean_url(vals[19]) if len(vals) > 19 else None,
            "linkedin": clean_url(vals[20]) if len(vals) > 20 else None,
            "tiktok": clean_url(vals[21]) if len(vals) > 21 else None,
            "website_org": clean_url(vals[22]) if len(vals) > 22 else None,
            "facebook_org": clean_url(vals[23]) if len(vals) > 23 else None,
            "twitter_org": clean_url(vals[25]) if len(vals) > 25 else None,
            "instagram_org": clean_url(vals[26]) if len(vals) > 26 else None,
            "linkedin_org": clean_url(vals[27]) if len(vals) > 27 else None,
        }

        if bday:
            rec["birthday"] = bday.isoformat()
            rec["hide_birthday_year"] = hide_year

        records[email] = {k: v for k, v in rec.items() if v is not None}

    wb.close()
    return records


def categorize_social_url(url, is_org=False):
    """Attempt to categorize a social URL by platform."""
    result = {}
    url_lower = url.lower()
    suffix = "_org" if is_org else ""

    if "linkedin.com" in url_lower:
        result[f"linkedin{suffix}" if is_org else "linkedin"] = url
    elif "twitter.com" in url_lower or "x.com" in url_lower:
        result[f"twitter{suffix}" if is_org else "twitter"] = url
    elif "instagram.com" in url_lower:
        result[f"instagram{suffix}" if is_org else "instagram"] = url
    elif "facebook.com" in url_lower or "fb.com" in url_lower:
        result[f"facebook{suffix}" if is_org else "facebook"] = url
    elif "tiktok.com" in url_lower:
        result["tiktok"] = url
    else:
        # Generic URL — treat as website
        result[f"website{suffix}" if is_org else "website"] = url

    return result


# ---------------------------------------------------------------------------
# Supabase Integration
# ---------------------------------------------------------------------------

class SupabaseClient:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.key = key
        self.headers = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
        }

    def get_fellows(self):
        """Fetch all fellows from Supabase."""
        url = f"{self.url}/rest/v1/fellows?select=*&order=id"
        resp = requests.get(url, headers=self.headers)
        if resp.status_code != 200:
            raise RuntimeError(f"Failed to fetch fellows: {resp.status_code} {resp.text}")
        return resp.json()

    def update_fellow(self, fellow_id, data):
        """Update a fellow record."""
        url = f"{self.url}/rest/v1/fellows?id=eq.{fellow_id}"
        headers = {**self.headers, "Prefer": "return=minimal"}
        resp = requests.patch(url, headers=headers, json=data)
        if resp.status_code not in (200, 204):
            raise RuntimeError(f"Failed to update fellow {fellow_id}: {resp.status_code} {resp.text}")


# ---------------------------------------------------------------------------
# Merge Logic
# ---------------------------------------------------------------------------

def merge_records(global_data, cpf_data, ggf_data):
    """
    Merge records from all three sources.
    Priority: GGF > CPF > Global Alumni (more recent/specific data wins).
    """
    merged = {}

    # Start with Global Alumni (broadest, least recent)
    for email, rec in global_data.items():
        merged[email] = dict(rec)

    # Layer on CPF data (more specific, current cohort)
    for email, rec in cpf_data.items():
        if email in merged:
            for k, v in rec.items():
                if k == "source":
                    merged[email].setdefault("sources", []).append(v)
                    continue
                # CPF data takes priority over Global Alumni for matching fields
                if v is not None:
                    merged[email][k] = v
        else:
            merged[email] = dict(rec)

    # Layer on GGF data (most specific, current cohort)
    for email, rec in ggf_data.items():
        if email in merged:
            for k, v in rec.items():
                if k == "source":
                    merged[email].setdefault("sources", []).append(v)
                    continue
                if v is not None:
                    merged[email][k] = v
        else:
            merged[email] = dict(rec)

    return merged


def compute_updates(fellow, enrichment):
    """
    Given an existing fellow record and enrichment data,
    compute which fields should be updated (only fill blanks).
    Returns dict of fields to update.
    """
    updates = {}

    for field in ENRICHABLE_FIELDS:
        if field == "hide_birthday_year":
            continue  # Handled with birthday
        if field not in enrichment:
            continue

        new_val = enrichment[field]
        if new_val is None:
            continue

        # Get current DB value
        current_val = fellow.get(field)

        # Only update if current value is empty/null
        if current_val is None or (isinstance(current_val, str) and current_val.strip() == ""):
            updates[field] = new_val

            # If updating birthday, also set hide_birthday_year
            if field == "birthday" and "hide_birthday_year" in enrichment:
                updates["hide_birthday_year"] = enrichment["hide_birthday_year"]

    return updates


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    apply_mode = "--apply" in sys.argv
    verbose = "--verbose" in sys.argv

    if apply_mode:
        print("=" * 60)
        print("  LIVE MODE — Changes will be written to Supabase")
        print("=" * 60)
    else:
        print("=" * 60)
        print("  DRY RUN — No changes will be made")
        print("  Use --apply to write changes to Supabase")
        print("=" * 60)

    print()

    # Check environment
    SUPABASE_URL = os.environ.get("SUPABASE_URL")
    SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        print("Error: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY environment variables required")
        print("Set them before running this script:")
        print("  export SUPABASE_URL=https://your-project.supabase.co")
        print("  export SUPABASE_SERVICE_ROLE_KEY=your-service-role-key")
        sys.exit(1)

    db = SupabaseClient(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)

    # Step 1: Parse all Excel files
    print("Step 1: Parsing historical Excel files...")
    print()

    print("  [1/3] Global Alumni Network Directory")
    global_data = parse_global_alumni()
    print(f"         → {len(global_data)} records extracted")
    print()

    print("  [2/3] 2025 CPF Tracker")
    cpf_data = parse_cpf_tracker()
    print(f"         → {len(cpf_data)} records extracted")
    print()

    print("  [3/3] 2025 GGF Tracker")
    ggf_data = parse_ggf_tracker()
    print(f"         → {len(ggf_data)} records extracted")
    print()

    # Step 2: Merge records
    print("Step 2: Merging records from all sources...")
    merged = merge_records(global_data, cpf_data, ggf_data)
    print(f"  → {len(merged)} unique email addresses with enrichment data")
    print()

    # Step 3: Fetch existing fellows from Supabase
    print("Step 3: Fetching existing fellows from Supabase...")
    fellows = db.get_fellows()
    print(f"  → {len(fellows)} fellows in database")

    # Build email → fellow lookup (including alternate_emails)
    email_to_fellow = {}
    for f in fellows:
        if f.get("email"):
            email_to_fellow[f["email"].lower()] = f
        for alt in (f.get("alternate_emails") or []):
            if alt:
                email_to_fellow[alt.lower()] = f

    print(f"  → {len(email_to_fellow)} email addresses indexed (including alternates)")
    print()

    # Step 4: Compute updates
    print("Step 4: Computing updates...")
    all_updates = []  # List of (fellow, updates_dict, matched_email)
    unmatched_emails = []
    field_counts = defaultdict(int)

    for email, enrichment in merged.items():
        fellow = email_to_fellow.get(email)
        if not fellow:
            unmatched_emails.append(email)
            continue

        updates = compute_updates(fellow, enrichment)
        if updates:
            all_updates.append((fellow, updates, email))
            for field in updates:
                field_counts[field] += 1

    print(f"  → {len(all_updates)} fellows will be updated")
    print(f"  → {len(unmatched_emails)} emails from Excel not matched in DB")
    print()

    # Step 5: Show summary
    print("=" * 60)
    print("  UPDATE SUMMARY")
    print("=" * 60)
    print()

    if field_counts:
        print("  Fields to be enriched:")
        for field, count in sorted(field_counts.items(), key=lambda x: -x[1]):
            label = ENRICHABLE_FIELDS.get(field, field)
            print(f"    {label:25s} → {count:3d} fellows")
        print()

    # Show detailed updates
    if verbose or not apply_mode:
        print("  Detailed changes per fellow:")
        print("  " + "-" * 56)

        for fellow, updates, email in sorted(all_updates, key=lambda x: x[0].get("last_name", "")):
            name = f"{fellow.get('first_name', '?')} {fellow.get('last_name', '?')}"
            print(f"\n  {name} ({fellow['id']}) — matched via {email}")
            for field, value in sorted(updates.items()):
                label = ENRICHABLE_FIELDS.get(field, field)
                current = fellow.get(field)
                display_val = str(value)[:60]
                if current:
                    print(f"    {label:22s}: {display_val}  (current: {str(current)[:30]})")
                else:
                    print(f"    {label:22s}: {display_val}")

        print()

    # Show unmatched emails (potential new contacts or typos)
    if unmatched_emails and verbose:
        print(f"  Unmatched emails ({len(unmatched_emails)}):")
        for email in sorted(unmatched_emails):
            src = merged[email].get("source", "?")
            name = f"{merged[email].get('first_name', '?')} {merged[email].get('last_name', '?')}"
            print(f"    {email:40s} — {name} ({src})")
        print()

    # Step 6: Apply updates
    if apply_mode and all_updates:
        print("Step 5: Applying updates to Supabase...")
        success = 0
        errors = 0

        for fellow, updates, email in all_updates:
            try:
                db.update_fellow(fellow["id"], updates)
                success += 1
                if verbose:
                    print(f"  Updated {fellow.get('first_name')} {fellow.get('last_name')} ({fellow['id']})")
            except Exception as e:
                errors += 1
                print(f"  ERROR updating {fellow['id']}: {e}")

        print()
        print(f"  Done! {success} fellows updated, {errors} errors")
    elif not apply_mode:
        print("  Dry run complete. Use --apply to write changes to Supabase.")
        print()
        # Print a quick command reminder
        print("  To apply:")
        print("    export SUPABASE_URL=https://pwazurumpzydxyppbvee.supabase.co")
        print("    export SUPABASE_SERVICE_ROLE_KEY=<your-service-role-key>")
        print("    python3 scripts/import-historical-data.py --apply")

    print()


if __name__ == "__main__":
    main()
